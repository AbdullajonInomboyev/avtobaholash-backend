"""
Excel Importer — Foydalanuvchilarni Excel fayldan import qilish
"""
import logging
from django.contrib.auth import get_user_model
from django.db import transaction

logger = logging.getLogger(__name__)
User = get_user_model()


class ExcelImporter:
    """
    Excel format (A ustun — Familiya, B — Ism, C — Email, D — Parol ixtiyoriy):
    Familiya | Ism | Otasining ismi | Email | Talaba ID | Guruh
    """

    REQUIRED_COLUMNS = ['familiya', 'ism', 'email']

    def __init__(self, log):
        self.log = log
        self.errors = []
        self.success_count = 0

    def run(self):
        from apps.question_bank.models import BulkImportLog
        try:
            rows = self._read_excel()
            self.log.total_rows = len(rows)
            self.log.save(update_fields=['total_rows'])

            for i, row in enumerate(rows, start=2):
                try:
                    # Har bir satr uchun alohida transaction — biri xato bo'lsa boshqalarga ta'sir qilmaydi
                    with transaction.atomic():
                        self._process_row(row, i)
                    self.success_count += 1
                except Exception as e:
                    self.errors.append({'row': i, 'error': str(e)})

            self.log.success_rows = self.success_count
            self.log.error_rows = len(self.errors)
            self.log.error_details = self.errors
            self.log.status = BulkImportLog.ImportStatus.DONE
            self.log.save()

        except Exception as e:
            logger.error(f'Import failed: {e}')
            self.log.status = BulkImportLog.ImportStatus.FAILED
            self.log.error_details = [{'row': 0, 'error': str(e)}]
            self.log.save()
            raise

    def _read_excel(self) -> list:
        import openpyxl
        wb = openpyxl.load_workbook(self.log.file.path)
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(dict(zip(headers, row)))

        return rows

    def _process_row(self, row: dict, row_num: int):
        email = str(row.get('email', '') or '').strip()
        first_name = str(row.get('ism', '') or '').strip()
        last_name = str(row.get('familiya', '') or '').strip()

        if not all([email, first_name, last_name]):
            raise ValueError('Familiya, Ism va Email to\'ldirilishi shart')

        if User.objects.filter(email=email).exists():
            raise ValueError(f'Bu email allaqachon mavjud: {email}')

        import secrets
        password = row.get('parol') or secrets.token_urlsafe(10)

        ROLE_MAP = {
            'students': 'student',
            'teachers': 'teacher',
        }
        role = ROLE_MAP.get(self.log.import_type)
        if not role:
            raise ValueError(f'Noto\'g\'ri import turi: {self.log.import_type}')

        user = User.objects.create_user(
            email=email,
            password=str(password),
            first_name=first_name,
            last_name=last_name,
            middle_name=str(row.get('otasining_ismi', '') or '').strip(),
            student_id=str(row.get('talaba_id', '') or '').strip(),
            role=role,
            tenant=self.log.tenant,
        )

        # Guruhga qo'shish
        group_name = str(row.get('guruh', '') or '').strip()
        if group_name and self.log.import_type == 'students':
            self._assign_to_group(user, group_name)

    def _assign_to_group(self, user, group_name: str):
        from apps.organization.models import Group, StudentGroup
        try:
            group = Group.objects.get(
                tenant=self.log.tenant,
                name__iexact=group_name,
                is_active=True,
            )
            StudentGroup.objects.get_or_create(
                tenant=self.log.tenant,
                student=user,
                group=group,
            )
        except Group.DoesNotExist:
            logger.warning(f'Guruh topilmadi: {group_name}')
